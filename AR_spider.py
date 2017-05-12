#!/usr/bin/env python
#encoding=UTF-8
#lxml required
#try:
#   cd C:\Python27\Scripts
#   easy_install.exe lxml
from openpyxl import  load_workbook
import lxml.etree
import lxml.html
import urllib2
import re
import time
import platform

def balabala(code):
    print "Code:", code
    main_doc=lxml.etree.HTML(urllib2.urlopen("http://vip.stock.finance.sina.com.cn/corp/go.php/vCB_Bulletin/stockid/{code}/page_type/ndbg.phtml".format(code=code)).read())
    code, city = re.search(r"\((\d+)\.(\w+)\)", str(iter(main_doc.xpath("//h1[@id='stockName']/span/text()")).next())).groups()
    name = iter(main_doc.xpath("//h1[@id='stockName']/text()")).next()

    dates = [time.strptime(e.strip(), "%Y-%m-%d")
        for e in main_doc.xpath("//div[@class='datelist']/ul/text()")
        if e.strip()]

    txtes = ["annual_report/{city}{code}_{year}.txt".format(
                city = city,
                code = code,
                year = d.tm_year - 1,
            ) for d in dates]

    urls = main_doc.xpath("//div[@class='datelist']/ul/a/@href")

    for _, url, filename in zip(dates,urls,txtes):
        print "Downloading:", filename
        open(filename, 'a').write(re.search(r'<pre>([\s\S\n]+)</pre>', urllib2.urlopen("http://vip.stock.finance.sina.com.cn" + url).read()).group(1))

    return '\n'.join([ '\t'.join(map(unicode, [name, code, u, t]))
        for d, u, t in zip(dates,urls,txtes)])

def get_num():
    num_ = []
    workbook_ = load_workbook(filename='stockNum.xlsx', read_only=True)
    sheetnames = workbook_.get_sheet_names()
    sheet = workbook_.get_sheet_by_name(sheetnames[0])
    os = platform.system() 
    if os == 'Linux':
        rows = len(sheet.rows)
        c = 0
    elif os == 'Windows':
        ws = workbook_[sheetnames[0]]
        rows = sum(1 for i in ws.rows)
        c = 1
    print(rows)
    for rowNum in range(2, rows):
        num = sheet.cell(row=rowNum, column=c).value
        # print type(num),num
        num_.append(num)
    return num_

def girigiri(x):
  try:
    return balabala(x)
  except Exception:
    return ""

if __name__ == "__main__":
    num_ = get_num()
    open("index.tsv", 'w').write("\t".join(["name", "code", "url", "filename"]) + "\n" + "\n".join(map(girigiri,
        num_,
    )).encode("utf-8"))


